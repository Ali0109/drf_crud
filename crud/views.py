from django.http import Http404
from rest_framework import status, permissions
from rest_framework.authentication import SessionAuthentication, BasicAuthentication, TokenAuthentication
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticatedOrReadOnly, IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from django.utils.crypto import get_random_string

from openpyxl import Workbook

from crud.models import *
from crud.serializers import *


class ExcelAPIView(APIView):
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws1_name = "Mysheet"
        ws1 = wb.create_sheet(ws1_name)
        ws2_name = "Mysheet0"
        ws2 = wb.create_sheet("Mysheet0", 0)
        ws3_name = "Mysheet-1"
        ws3 = wb.create_sheet("Mysheet-1", -1)
        for i in range(1, 20):
            ws[f'A{i}'] = f"Something-{i}"
            ws1[f'A{i}'] = f"{ws1_name}-{i}"
            ws2[f'A{i}'] = f"{ws2_name}-{i}"
            ws3[f'A{i}'] = f"{ws3_name}-{i}"
        unique_name = get_random_string(length=32)
        wb.save(f"media/csv/{unique_name}.xlsx")
        return Response("OK")


class DataAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        all_data = Data.objects.all()
        serializer = DataSerializer(all_data, many=True)
        return Response({"message": serializer.data})

    def post(self, request):
        serializer = DataSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DataDetailAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def get_object(self, pk):
        try:
            return Data.objects.get(pk=pk)
        except Data.DoesNotExist:
            raise Http404

    def get(self, request, pk):
        pk_data = self.get_object(pk)
        serializer = DataSerializer(pk_data, many=False)
        return Response({"message": serializer.data})

    def put(self, request, pk):
        pk_data = self.get_object(pk)
        serializer = DataSerializer(pk_data, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": serializer.data})
        return Response({"message": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        pk_data = self.get_object(pk)
        pk_data.delete()
        return Response({'message': "Data was deleted successfully"}, status=status.HTTP_204_NO_CONTENT)


@api_view(["PUT"])
@permission_classes(IsAuthenticated,)
def DataRestoreDetailAPIView(request, pk):
    if request.method == "PUT":
        try:
            pk_data = Data.deleted_objects.get(pk=pk)
            pk_data.restore()
            return Response({"message": "Data was restore successfully"})
        except Data.DoesNotExist:
            raise Http404

